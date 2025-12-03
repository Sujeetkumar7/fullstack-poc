
import sys
from awsglue.transforms import *
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from awsglue.context import GlueContext
from awsglue.job import Job
from pyspark.sql import functions as F
from datetime import datetime
import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Get job arguments
args = getResolvedOptions(sys.argv, ['JOB_NAME'])
sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args['JOB_NAME'], args)

# S3 details
bucket_name = "wmnanalytics"
folder_prefix = "analytics-output"
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_key = f"{folder_prefix}/Combined_Analytics_{timestamp}.xlsx"
local_path = f"/tmp/Analytics_Report.xlsx"

# Read data from Glue Catalog
dyf = glueContext.create_dynamic_frame.from_catalog(database="wmndb", table_name="transaction")
df = dyf.toDF().cache()

# Convert amount to double and handle nulls
df = df.withColumn("amount_double", F.col("amount").cast("double")).na.fill({"amount_double": 0})

# Analytics 1: User Summary
user_summary = (
    df.groupBy("user_id")
      .agg(
          F.max("amount_double").alias("Max Amount"),
          F.min("amount_double").alias("Min Amount"),
          F.avg("amount_double").alias("Average Amount"),
          F.sum("amount_double").alias("Total Amount"),
          F.count("*").alias("Transaction Count")
      )
      .withColumnRenamed("user_id", "User ID")
)

# Analytics 2: Daily Trend
daily_trend = (
    df.withColumn("Date", F.to_date("timestamp"))
      .groupBy("Date")
      .agg(F.sum("amount_double").alias("Total Amount"), F.count("*").alias("Transaction Count"))
)

# Analytics 3: Top Users
top_users = user_summary.orderBy(F.desc("Total Amount")).limit(10)

# Analytics 4: Type Summary
type_summary = (
    df.groupBy("transaction_type")
      .agg(F.sum("amount_double").alias("Total Amount"), F.count("*").alias("Count"))
      .withColumnRenamed("transaction_type", "Transaction Type")
)

# Convert to Pandas
user_summary_pd = user_summary.toPandas()
daily_trend_pd = daily_trend.toPandas()
top_users_pd = top_users.toPandas()
type_summary_pd = type_summary.toPandas()

# Write to Excel with multiple sheets
with pd.ExcelWriter(local_path, engine='openpyxl') as writer:
    user_summary_pd.to_excel(writer, sheet_name='User_Summary', index=False)
    daily_trend_pd.to_excel(writer, sheet_name='Daily_Trend', index=False)
    top_users_pd.to_excel(writer, sheet_name='Top_Users', index=False)
    type_summary_pd.to_excel(writer, sheet_name='Type_Summary', index=False)

# Apply styling
wb = load_workbook(local_path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
wb.save(local_path)

# Upload to S3
s3 = boto3.client('s3')
s3.upload_file(local_path, bucket_name, excel_key)
print(f"Excel report uploaded to s3://{bucket_name}/{excel_key}")

job.commit()
